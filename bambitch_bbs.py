#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BAMBITCH宇都宮 ご来店予告掲示板(KENT-WEB imgboard.cgi) アーカイバ。

公開掲示板の投稿を取得し、ひとつの Excel ワークシート bbs_log.xlsx に
「日付・投稿者名」を中心とした一覧として蓄積する。添付画像は img-box/ に保存。

- bbs_log.xlsx (シート "posts"): No. / 日付 / 時刻 / 名前(投稿者) / 題名 / 本文 / 画像ファイル
- img-box/imgYYYYMMDDHHMMSS.jpg: 添付画像(掲示板と同じファイル名)

重複は投稿 No. で排除するので、毎日実行しても新規分だけが追記される(冪等)。
xlsx 自体を一覧の唯一の保存先(source of truth)とし、既存 No. も xlsx から読む。

環境変数:
  BBS_MAX_PAGES  取得するページ数(既定: 3)
  BBS_BASE_URL   掲示板 CGI の URL(既定は BAMBITCH宇都宮)
"""
from __future__ import annotations

import html as html_lib
import os
import re
import sys
import time
from datetime import datetime, timezone, timedelta
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook

JST = timezone(timedelta(hours=9))

BASE_URL = os.environ.get(
    "BBS_BASE_URL",
    "https://bambitch-utsunomiya.com/cgi-bin-utnmy/imgboard.cgi",
)
# CGI と同じディレクトリ(相対 ./img-box/ の解決に使う)
ROOT_URL = BASE_URL.rsplit("/", 1)[0] + "/"

MAX_PAGES = int(os.environ.get("BBS_MAX_PAGES", "3"))

REPO_DIR = Path(__file__).resolve().parent
IMG_DIR = REPO_DIR / "img-box"
XLSX_PATH = REPO_DIR / "bbs_log.xlsx"
SHEET_NAME = "posts"
HEADERS = ["No.", "日付", "時刻", "名前", "出勤者", "題名", "本文", "画像ファイル"]

# お店の公式名義。原則この名義の投稿は記録しないが、出勤者紹介(本文に「担当」)だけ残す。
STORE_NAME = "★BAMBITCH★"

USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Safari/605.1.15"
)

session = requests.Session()
session.headers.update({"User-Agent": USER_AGENT})


def log(msg: str) -> None:
    print(f"[{datetime.now(JST):%Y-%m-%d %H:%M:%S}] {msg}", flush=True)


def fetch_page(page: int) -> str:
    """指定ページの HTML を Shift_JIS からデコードして返す。"""
    if page <= 1:
        url = BASE_URL
    else:
        url = f"{BASE_URL}?amode=&p1=&p2=&bbsaction=page_change&page={page}"
    resp = session.get(url, timeout=30)
    resp.raise_for_status()
    resp.encoding = "shift_jis"  # KENT-WEB imgboard は Shift_JIS
    return resp.text


def strip_tags(fragment: str) -> str:
    """<BR> を改行に、その他タグを除去してテキスト化する。"""
    text = re.sub(r"(?i)<br\s*/?>", "\n", fragment)
    text = re.sub(r"(?is)<[^>]+>", "", text)
    text = html_lib.unescape(text)
    lines = [ln.rstrip() for ln in text.splitlines()]
    return "\n".join(lines).strip()


# 各記事は削除用チェックボックス <INPUT ... NAME="rmid...S{timestamp}"> で始まる
ARTICLE_SPLIT = re.compile(r'(?i)<INPUT\s+TYPE="?CHECKBOX"?\s+NAME="rmid')

RE_NO = re.compile(r"No\.(\d+)")
RE_DATE = re.compile(r"\[(\d{4}/\d{2}/\d{2}),(\d{2}:\d{2}:\d{2})\]")
RE_IMG = re.compile(r'(?i)<IMG\s+SRC="(\.?/?img-box/[^"]+)"')
# 題名(subject): 大きめ FONT SIZE="+1" の太字
RE_TITLE = re.compile(r'(?i)<FONT\s+SIZE="?\+1"?[^>]*>\s*<B>(.*?)</B>', re.S)
# 投稿者名: 日付ブロック [YYYY/... の直前に置かれる太字。親記事は題名の次の FONT、
# 返信記事は "名前：" の後の FONT に入る。いずれも日付直前の <B>…</B></FONT> を取る。
RE_NAME = re.compile(r"(?is)<B>([^<]*)</B>\s*</FONT>\s*\[\d{4}/")
RE_BODY = re.compile(r"(?is)<BLOCKQUOTE[^>]*>(.*?)</BLOCKQUOTE>")
# 営業案内の出勤者紹介ブロック: 「担当は … なのよ」の間に1行1名で並ぶ
RE_STAFF = re.compile(r"担当は(.*?)なのよ", re.S)


def extract_staff(body: str) -> str:
    """本文の「担当は…なのよ」から出勤者の行を取り出し ' / ' 連結で返す。"""
    m = RE_STAFF.search(body)
    region = m.group(1) if m else ""
    lines = [ln.strip() for ln in region.splitlines() if ln.strip()]
    return " / ".join(lines)


def parse_articles(page_html: str) -> list[dict]:
    """1 ページ分の HTML から記事レコードのリストを返す。"""
    chunks = ARTICLE_SPLIT.split(page_html)
    articles: list[dict] = []
    for chunk in chunks[1:]:  # 先頭はヘッダ部
        m_no = RE_NO.search(chunk)
        m_date = RE_DATE.search(chunk)
        if not m_no or not m_date:
            continue
        # 返信記事(お店の返信など)はスキップ。KENT-WEB は返信を専用テンプレート
        # (コメント "返信用" / 名前欄に "名前：" ラベル)で出力するので、それを検出する。
        if "返信用" in chunk or "名前：" in chunk:
            continue
        no = int(m_no.group(1))
        date_str = f"{m_date.group(1)} {m_date.group(2)}"
        try:
            dt = datetime.strptime(date_str, "%Y/%m/%d %H:%M:%S")
        except ValueError:
            continue

        img_filename = None
        img_url = None
        m_img = RE_IMG.search(chunk)
        if m_img:
            rel = m_img.group(1).lstrip("./")  # img-box/imgXXXX.jpg
            img_url = ROOT_URL + rel
            img_filename = rel.split("/")[-1]

        m_title = RE_TITLE.search(chunk)
        title = strip_tags(m_title.group(1)) if m_title else ""

        m_name = RE_NAME.search(chunk)
        name = strip_tags(m_name.group(1)) if m_name else ""

        m_body = RE_BODY.search(chunk)
        body = strip_tags(m_body.group(1)) if m_body else ""

        staff = extract_staff(body)

        # お店(★BAMBITCH★)名義の投稿は原則除外。ただし出勤者紹介
        # (本文に「担当」= 出勤者あり)は記録として残す。来店客の投稿は対象外なので影響しない。
        if name == STORE_NAME and "担当" not in body:
            continue

        articles.append(
            {
                "no": no,
                "date": dt.strftime("%Y-%m-%d"),
                "time": dt.strftime("%H:%M:%S"),
                "name": name,
                "staff": staff,
                "title": title,
                "body": body,
                "image_filename": img_filename,
                "image_url": img_url,
            }
        )
    return articles


def download_image(img_url: str, filename: str) -> bool:
    """画像を img-box/ に保存。既存ならスキップ。"""
    dest = IMG_DIR / filename
    if dest.exists() and dest.stat().st_size > 0:
        return False
    try:
        resp = session.get(img_url, timeout=60)
        resp.raise_for_status()
        IMG_DIR.mkdir(parents=True, exist_ok=True)
        dest.write_bytes(resp.content)
        log(f"  画像取得: {filename} ({len(resp.content)//1024} KB)")
        return True
    except requests.RequestException as e:
        log(f"  画像取得失敗: {filename} ({e})")
        return False


# 名前欄の複数名の区切りは読点「、」とカンマ「,」「，」のみ。
# 「・」(例: オ・スー)は 1 名の一部なので分割しない。
NAME_SPLIT = re.compile(r"[、,，]")
VISITS_SHEET = "来店履歴"
RANK_SHEET = "来店回数"
WEEKDAY_SHEET = "曜日別"
HOUR_SHEET = "時間帯別"
MONTH_SHEET = "月次推移"
NETWORK_SHEET = "同伴ネットワーク"
DERIVED_SHEETS = (
    VISITS_SHEET, RANK_SHEET, WEEKDAY_SHEET, HOUR_SHEET, MONTH_SHEET, NETWORK_SHEET,
)
JP_WEEKDAYS = ["月", "火", "水", "木", "金", "土", "日"]


def split_names(name: str) -> list[str]:
    """名前欄を個人名に分割する(空白のみは除く)。"""
    return [p.strip() for p in NAME_SPLIT.split(name or "") if p.strip()]


def weekday_jp(date: str) -> str:
    """'YYYY-MM-DD' → 曜日(月〜日)。パースできなければ空。"""
    try:
        return JP_WEEKDAYS[datetime.strptime(date, "%Y-%m-%d").weekday()]
    except (ValueError, TypeError):
        return ""


def rebuild_visit_sheets(wb) -> None:
    """posts シートから、個人単位の各集計シートを作り直す。

    - 来店履歴: 1 人 1 行に展開。何回目 = その人の累積来店回数(No. 昇順=時系列)。
    - 来店回数: 個人ごとの合計来店回数・初回・最終日。回数の多い順。
    - 曜日別 / 時間帯別 / 月次推移: 投稿件数・延べ来店人数などの集計。
    - 同伴ネットワーク: 一緒に来た人物ペアの共起回数(エッジ一覧)。
    - お店(★BAMBITCH★)名義は来店客ではないので集計対象外。
    """
    from collections import Counter
    from itertools import combinations

    ws = wb[SHEET_NAME]
    posts = []
    for no, date, time_, name, staff, title, body, img in ws.iter_rows(
        min_row=2, values_only=True
    ):
        if no is None or name == STORE_NAME:
            continue
        posts.append((int(no), date, time_, name, title, body))
    posts.sort(key=lambda x: x[0])  # No. 昇順 = 時系列

    for sn in DERIVED_SHEETS:
        if sn in wb.sheetnames:
            del wb[sn]

    vh = wb.create_sheet(VISITS_SHEET)
    vh.append(["来店日", "曜日", "時刻", "No.", "個人名", "何回目", "同伴(元の名前)", "題名", "本文"])
    vh.freeze_panes = "A2"

    counter: dict[str, int] = {}
    totals: dict[str, int] = {}
    first_seen: dict[str, str] = {}
    last_seen: dict[str, str] = {}
    wd_visits: dict[str, int] = {d: 0 for d in JP_WEEKDAYS}
    wd_posts: dict[str, int] = {d: 0 for d in JP_WEEKDAYS}
    hr_visits = [0] * 24  # 時間帯(時)→延べ来店人数
    hr_posts = [0] * 24   # 時間帯(時)→投稿件数
    mon_posts: dict[str, int] = {}          # YYYY-MM→投稿件数
    mon_visits: dict[str, int] = {}         # YYYY-MM→延べ来店人数
    mon_people: dict[str, set] = {}         # YYYY-MM→ユニーク人物
    pair_counts: Counter = Counter()        # (A,B)→同伴回数

    for no, date, time_, name, title, body in posts:
        wd = weekday_jp(date)
        hour = int(time_[:2]) if isinstance(time_, str) and time_[:2].isdigit() else None
        month = date[:7] if isinstance(date, str) and len(date) >= 7 else ""
        people = split_names(name)

        if wd:
            wd_posts[wd] += 1
        if hour is not None:
            hr_posts[hour] += 1
        if month:
            mon_posts[month] = mon_posts.get(month, 0) + 1
            mon_people.setdefault(month, set())

        for person in people:
            counter[person] = counter.get(person, 0) + 1
            totals[person] = counter[person]
            first_seen.setdefault(person, date)
            last_seen[person] = date
            if wd:
                wd_visits[wd] += 1
            if hour is not None:
                hr_visits[hour] += 1
            if month:
                mon_visits[month] = mon_visits.get(month, 0) + 1
                mon_people[month].add(person)
            vh.append([date, wd, time_, no, person, counter[person], name, title, body])

        # 同伴ペア(同一投稿に2名以上): 名前順に正規化して共起カウント
        for a, b in combinations(sorted(set(people)), 2):
            pair_counts[(a, b)] += 1

    rk = wb.create_sheet(RANK_SHEET)
    rk.append(["個人名", "来店回数", "初回", "最終"])
    rk.freeze_panes = "A2"
    for person in sorted(totals, key=lambda p: (-totals[p], p)):
        rk.append([person, totals[person], first_seen[person], last_seen[person]])

    wk = wb.create_sheet(WEEKDAY_SHEET)
    wk.append(["曜日", "投稿件数", "延べ来店人数"])
    wk.freeze_panes = "A2"
    for d in JP_WEEKDAYS:
        wk.append([d, wd_posts[d], wd_visits[d]])

    hs = wb.create_sheet(HOUR_SHEET)
    hs.append(["時間帯", "投稿件数", "延べ来店人数"])
    hs.freeze_panes = "A2"
    for h in range(24):
        hs.append([f"{h:02d}時台", hr_posts[h], hr_visits[h]])

    ms = wb.create_sheet(MONTH_SHEET)
    ms.append(["月", "投稿件数", "延べ来店人数", "ユニーク人数"])
    ms.freeze_panes = "A2"
    for month in sorted(mon_posts):
        ms.append([month, mon_posts[month], mon_visits.get(month, 0), len(mon_people[month])])

    nw = wb.create_sheet(NETWORK_SHEET)
    nw.append(["人物A", "人物B", "同伴回数"])
    nw.freeze_panes = "A2"
    for (a, b), c in sorted(pair_counts.items(), key=lambda kv: (-kv[1], kv[0])):
        nw.append([a, b, c])


def load_workbook_or_new():
    """既存の bbs_log.xlsx を開く。無ければ見出し付きで新規作成する。"""
    if XLSX_PATH.exists():
        wb = load_workbook(XLSX_PATH)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        return wb, ws
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    ws.freeze_panes = "A2"  # 見出し行を固定
    return wb, ws


def existing_nos(ws) -> set[str]:
    """シート内の既存 No.(A 列)を集合で返す。"""
    nos = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row and row[0] is not None:
            nos.add(str(row[0]))
    return nos


def main() -> int:
    log(f"取得開始: {BASE_URL} (最大 {MAX_PAGES} ページ)")
    wb, ws = load_workbook_or_new()
    seen = existing_nos(ws)

    all_articles: list[dict] = []
    for page in range(1, MAX_PAGES + 1):
        try:
            page_html = fetch_page(page)
        except requests.RequestException as e:
            log(f"ページ {page} 取得失敗: {e}")
            break
        arts = parse_articles(page_html)
        log(f"ページ {page}: {len(arts)} 件の記事を検出")
        if not arts:
            break
        all_articles.extend(arts)
        time.sleep(1.5)  # サーバ負荷に配慮

    # No. で重複排除(ページ跨ぎ)し、古い順に追記
    unique: dict[int, dict] = {}
    for a in all_articles:
        unique[a["no"]] = a

    new_posts = [unique[no] for no in sorted(unique) if str(no) not in seen]

    for p in new_posts:
        if p["image_url"] and p["image_filename"]:
            download_image(p["image_url"], p["image_filename"])
        ws.append(
            [
                p["no"],
                p["date"],
                p["time"],
                p["name"],
                p["staff"],
                p["title"],
                p["body"],
                p["image_filename"] or "",
            ]
        )

    if new_posts:
        rebuild_visit_sheets(wb)  # 来店履歴・来店回数を最新化
        wb.save(XLSX_PATH)
        log(f"新規投稿 {len(new_posts)} 件をワークシートに追加しました。")
        for p in new_posts:
            log(f"  + No.{p['no']} [{p['date']} {p['time']}] {p['name']} / {p['title'][:24]}")
    else:
        log("新規投稿はありませんでした。")

    total = ws.max_row - 1  # 見出しを除く
    log(f"完了。総投稿数(累計): {total}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
