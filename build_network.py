#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""同伴ネットワークを力学グラフの HTML(単体ファイル)として書き出す。

bbs_log.xlsx の「同伴ネットワーク」(エッジ)と「来店回数」(ノードの重み)を読み、
外部ライブラリ不要の自己完結 HTML(network.html)を生成する。ブラウザで開くと
ノード=人物(大きさ=来店回数)、線=一緒に来た関係(太さ=同伴回数)が表示される。

使い方:
  python build_network.py            # network.html を出力
  python build_network.py out.html   # 出力先を指定
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

REPO_DIR = Path(__file__).resolve().parent
XLSX_PATH = REPO_DIR / "bbs_log.xlsx"


def build(out_path: Path) -> None:
    wb = load_workbook(XLSX_PATH, read_only=True)
    edges = []
    node_ids = set()
    for a, b, c in wb["同伴ネットワーク"].iter_rows(min_row=2, values_only=True):
        if a is None or b is None:
            continue
        edges.append({"source": a, "target": b, "weight": int(c)})
        node_ids.update((a, b))

    visits = {}
    for person, cnt, *_ in wb["来店回数"].iter_rows(min_row=2, values_only=True):
        if person is not None:
            visits[person] = int(cnt)

    nodes = [{"id": p, "visits": visits.get(p, 1)} for p in sorted(node_ids)]

    html = _TEMPLATE.replace("__NODES__", json.dumps(nodes, ensure_ascii=False)).replace(
        "__EDGES__", json.dumps(edges, ensure_ascii=False)
    )
    out_path.write_text(html, encoding="utf-8")
    print(f"出力: {out_path}  (ノード {len(nodes)} / エッジ {len(edges)})")


_TEMPLATE = r"""<!doctype html>
<html lang="ja"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>BAMBITCH宇都宮 同伴ネットワーク</title>
<style>
  html,body{margin:0;height:100%;background:#0f1117;color:#e6e6e6;
    font-family:-apple-system,"Hiragino Kaku Gothic ProN",Meiryo,sans-serif;overflow:hidden}
  #hud{position:fixed;top:10px;left:12px;font-size:13px;line-height:1.6;
    background:rgba(20,24,34,.8);padding:8px 12px;border-radius:8px;pointer-events:none}
  #hud b{color:#ffd34d}
  svg{width:100vw;height:100vh;display:block}
  line{stroke:#4a5570;stroke-opacity:.5}
  circle{fill:#3aa0ff;stroke:#0f1117;stroke-width:1.5;cursor:grab}
  text{fill:#cfd6e4;font-size:11px;pointer-events:none;text-shadow:0 1px 2px #000}
</style></head><body>
<div id="hud"><b>同伴ネットワーク</b> — ノード=人物(大=来店多)/線=一緒に来た関係<br>
ドラッグで移動・ホイールで拡大縮小</div>
<svg id="svg"><g id="view"></g></svg>
<script>
const nodes=__NODES__, links=__EDGES__;
const svg=document.getElementById('svg'), view=document.getElementById('view');
let W=innerWidth,H=innerHeight;
const byId=new Map(nodes.map(n=>[n.id,n]));
nodes.forEach(n=>{n.x=W/2+(Math.random()-.5)*W*.6;n.y=H/2+(Math.random()-.5)*H*.6;n.vx=0;n.vy=0;
  n.r=6+Math.sqrt(n.visits)*3;});
links.forEach(l=>{l.s=byId.get(l.source);l.t=byId.get(l.target);});
// build SVG
const NS='http://www.w3.org/2000/svg';
const lel=links.map(l=>{const e=document.createElementNS(NS,'line');
  e.setAttribute('stroke-width',Math.min(1+l.weight,5));view.appendChild(e);return e;});
const gnodes=nodes.map(n=>{const g=document.createElementNS(NS,'g');
  const c=document.createElementNS(NS,'circle');c.setAttribute('r',n.r);
  const t=document.createElementNS(NS,'text');t.setAttribute('x',n.r+2);t.setAttribute('y',4);
  t.textContent=n.id;g.appendChild(c);g.appendChild(t);view.appendChild(g);
  c.addEventListener('pointerdown',ev=>{drag=n;n.fx=n.x;n.fy=n.y;ev.target.setPointerCapture(ev.pointerId);});
  return {n,g,c};});
// force simulation
let drag=null, alpha=1;
function tick(){
  alpha*=0.99; if(alpha<0.005)alpha=0.005;
  // repulsion
  for(let i=0;i<nodes.length;i++)for(let j=i+1;j<nodes.length;j++){
    const a=nodes[i],b=nodes[j];let dx=a.x-b.x,dy=a.y-b.y;let d2=dx*dx+dy*dy||1;
    let f=1200/d2;let d=Math.sqrt(d2);dx/=d;dy/=d;
    a.vx+=dx*f;a.vy+=dy*f;b.vx-=dx*f;b.vy-=dy*f;}
  // springs
  links.forEach(l=>{let dx=l.t.x-l.s.x,dy=l.t.y-l.s.y;let d=Math.sqrt(dx*dx+dy*dy)||1;
    let f=(d-70)*0.02;dx/=d;dy/=d;
    l.s.vx+=dx*f;l.s.vy+=dy*f;l.t.vx-=dx*f;l.t.vy-=dy*f;});
  // gravity to center
  nodes.forEach(n=>{n.vx+=(W/2-n.x)*0.002;n.vy+=(H/2-n.y)*0.002;
    n.x+=n.vx*alpha;n.y+=n.vy*alpha;n.vx*=0.85;n.vy*=0.85;});
  if(drag){drag.x=drag.fx;drag.y=drag.fy;}
  lel.forEach((e,i)=>{const l=links[i];e.setAttribute('x1',l.s.x);e.setAttribute('y1',l.s.y);
    e.setAttribute('x2',l.t.x);e.setAttribute('y2',l.t.y);});
  gnodes.forEach(o=>o.g.setAttribute('transform',`translate(${o.n.x},${o.n.y})`));
  requestAnimationFrame(tick);
}
tick();
// drag + pan/zoom
addEventListener('pointermove',ev=>{if(drag){const p=toView(ev);drag.fx=p.x;drag.fy=p.y;alpha=0.3;}});
addEventListener('pointerup',()=>{if(drag){drag.fx=drag.fy=null;drag=null;}});
let scale=1,tx=0,ty=0;
function apply(){view.setAttribute('transform',`translate(${tx},${ty}) scale(${scale})`);}
function toView(ev){return {x:(ev.clientX-tx)/scale,y:(ev.clientY-ty)/scale};}
addEventListener('wheel',ev=>{ev.preventDefault();const k=ev.deltaY<0?1.1:0.9;
  tx=ev.clientX-(ev.clientX-tx)*k;ty=ev.clientY-(ev.clientY-ty)*k;scale*=k;apply();},{passive:false});
addEventListener('resize',()=>{W=innerWidth;H=innerHeight;});
apply();
</script></body></html>"""


def main() -> int:
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else REPO_DIR / "network.html"
    build(out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
