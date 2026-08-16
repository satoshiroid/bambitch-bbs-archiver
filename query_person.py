#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""特定の人物の来店履歴を bbs_log.xlsx から抜き出す。

使い方:
  python query_person.py ゆずき          # 完全一致
  python query_person.py -c ゆず          # 部分一致(名前に含む)
  python query_person.py ゆずき --csv out.csv   # CSV にも書き出す

集計元は「来店履歴」シート(bambitch_bbs.py が生成)。
"""
from __future__ import annotations

import argparse
import csv
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

XLSX_PATH = Path(__file__).resolve().parent / "bbs_log.xlsx"
VISITS_SHEET = "来店履歴"
JP_WEEKDAYS = ["月", "火", "水", "木", "金", "土", "日"]
# 来店履歴の列: 来店日, 曜日, 時刻, No., 個人名, 何回目, 同伴, 題名, 本文
C_DATE, C_WD, C_TIME, C_NO, C_PERSON, C_NTH, C_GROUP, C_TITLE, C_BODY = range(9)


def load_rows() -> list[tuple]:
    wb = load_workbook(XLSX_PATH, read_only=True)
    ws = wb[VISITS_SHEET]
    return [r for r in ws.iter_rows(min_row=2, values_only=True) if r[C_NO] is not None]


def main() -> int:
    ap = argparse.ArgumentParser(description="特定人物の来店履歴を抽出")
    ap.add_argument("name", help="対象の個人名")
    ap.add_argument("-c", "--contains", action="store_true", help="部分一致で検索")
    ap.add_argument("--csv", metavar="PATH", help="結果を CSV に書き出す")
    args = ap.parse_args()

    rows = load_rows()

    def match(person: str) -> bool:
        person = person or ""
        return args.name in person if args.contains else person == args.name

    hits = [r for r in rows if match(r[C_PERSON])]
    if not hits:
        mode = "を含む" if args.contains else "に一致する"
        print(f"『{args.name}』{mode}来店記録は見つかりませんでした。")
        return 1

    people = sorted(set(r[C_PERSON] for r in hits))
    for person in people:
        ph = sorted((r for r in hits if r[C_PERSON] == person), key=lambda r: r[C_NO])
        print(f"\n■ {person} — 来店 {len(ph)} 回")
        wd = Counter(r[C_WD] for r in ph)
        print("  曜日別: " + "  ".join(f"{d}:{wd.get(d, 0)}" for d in JP_WEEKDAYS))
        print("  " + "-" * 60)
        for r in ph:
            body = (r[C_BODY] or "").replace("\n", " ")[:28]
            group = r[C_GROUP] if r[C_GROUP] != person else "(単独)"
            print(
                f"  {r[C_DATE]}({r[C_WD]}) {r[C_TIME]}  No.{r[C_NO]}  "
                f"{r[C_NTH]}回目  題名={r[C_TITLE]}  同伴={group}  「{body}」"
            )

    if args.csv:
        with open(args.csv, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["来店日", "曜日", "時刻", "No.", "個人名", "何回目", "同伴", "題名", "本文"])
            for r in sorted(hits, key=lambda r: (r[C_PERSON], r[C_NO])):
                w.writerow(r)
        print(f"\nCSV を書き出しました: {args.csv}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
